"""Product Catalogue Updater

This script requests end-of-life data from the Cisco EoX API and updates the 'Product Catalogue.xlsx' spreadsheet
with the relevant dates.

The script runs as follows:

    *The existence and validity of an access token is checked, and a new one requested if necessary (eg. if it has
     expired). An access token is required for all calls to the Cisco EoX API.
    *The 'Product Catalogue.xlsx' spreadsheet is opened as a 'workbook' object with the 'openpyxl' module.
    *The hardware sheet of the Product Catalogue is selected, and a list of the Product ID's in the 'PartNumber' column
     is created.
    *For each product ID, the relevant url is created and a GET request sent to the API.
    *If an error is returned, the error message is printed and the process continues to the next Product ID.
    *The response is saved as a dict, from which the 'EOXRecord' is extracted.
    *The two relevant dates are written to the correct cells in the workbook, and a timestamp is saved.
    *If a response is not present for one of the dates, the value in the spreadsheet is not altered.
    *A separate text report for the hardware and software updates is saved, and the updated spreadsheet is saved.

This process is then repeated for the software sheet of the Product Catalogue.

The Cisco EoX API provides access to Cisco End of Life product data and is the source of information used to update
the Product Catalogue in this program. Calls to the API require different forms depending on what data is being
requested. For example, one method is to request data by Product ID; this method is used for the Hardware section
of this program. Documentation, support, and examples of using the Cisco Eox API can be found at the following location:

    *https://developer.cisco.com/docs/support-apis/#!eox

This script uses the 3rd party modules 'requests_oauthlib' and 'oauthlib' to handle the OAuth 2.0 security
authorisation required by the Cisco EoX API. Documentation for these modules can be found at the below addresses:

    *https://github.com/requests/requests-oauthlib
    *https://oauthlib.readthedocs.io/en/latest/index.html

The tool requires that a file 'Product Catalogue.xlsx' is present in the active directory. A JSON file named
'access_token.json' will be created to store the API access token. If this file already exists, it will be
over-written each time the program runs.

The program will output a file named 'Product Catalogue Updated **-**-**.xlsx', containing the date in the file name.
It will also output two text files, 'HW Report ****-**-**.txt' and 'SW Report ****-**-**.txt', where the asterisks
represent today's date. These text files contain a report on any changes that were made to the Product Catalogue.

The script requires that the following modules are installed within the active Python environment:

    *requests
    *requests_oauthlib
    *oauthlib
    *openpyxl
"""

import json
import sys
import datetime
import time
import openpyxl
import requests
from datetime import datetime
from datetime import date
from requests_oauthlib import OAuth2Session
from requests.auth import HTTPBasicAuth
from oauthlib.oauth2 import BackendApplicationClient


# API credentials
client_id = 'mpfz8vkej89nma4h6sbscuv7'
client_secret = 'bbT7UjWpNycdR9AkD6AvrSCt'
grant_type = 'client_credentials'
auth_url = 'https://cloudsso.cisco.com/as/token.oauth2'
hw_base_url = 'https://api.cisco.com/supporttools/eox/rest/5/EOXByProductID/1'
sw_base_url = 'https://api.cisco.com/supporttools/eox/rest/5/EOXBySWReleaseString/1'

# file names
token_file_name = 'access_token.json'
excel_name = 'Product Catalogue.xlsx'
hw_sheet_name = 'HW Product Catalogue'
sw_sheet_name = 'SW Product Catalogue'

scan_limit = 5000  # number of products to update per sheet


def create_report(sheet_name, scanned_count, updated_count, list_of_updates, name):
    """save a report on the results of the script execution to a text file"""
    file_name = name + '.txt'
    try:
        file = open(file_name, 'w')
    except IOError:
        print('Could not open file: {}'.format(file_name))
        sys.exit()
    else:
        file.write("Products scanned on {}: {}\n".format(sheet_name, scanned_count))
        file.write("Products updated: {}\n\n".format(updated_count))
        for product in list_of_updates:
            file.write('{}\n'.format(product))

        file.close()


def get_token(client_id, client_secret, auth_url):
    """get API token object from the authorisation URL"""
    auth = HTTPBasicAuth(client_id, client_secret)  # basic auth header for token request
    client = BackendApplicationClient(client_id=client_id)
    oauth = OAuth2Session(client=client)
    try:
        token = oauth.fetch_token(token_url=auth_url, auth=auth)
    except requests.ConnectionError as e:
        print('Could not fetch access token, check VPN is disconnected')
        print('Error: {}'.format(e))
        input('Press enter to continue.')
        sys.exit()
    return token


def load_token(token_file):
    """check if token_file exists, read the contents if it does"""
    data = json.load(token_file)  # read token object data
    return data


def save_token(name, token):
    """save token data to JSON file"""
    json_data = json.dumps(token)
    try:
        file = open(name, 'w')
    except IOError:
        print('Could not open file: {}'.format(name))
        sys.exit()
    else:
        file.write(json_data)
        file.close()
        return


def check_token(name):
    """Check if a file containing the access token already exists.

    If it exists, read the access token from the file. If it does not exist or it has expired, request a new access
    token and save it to the file.
    """
    try:
        file = open(name, 'r')
    except FileNotFoundError:
        print('Access token not found, requesting new one.')
        token = get_token(client_id, client_secret, auth_url)
        save_token(token_file_name, token)  # save to file for next time
    else:
        token = load_token(file)
        # check if token has expired
        if token['expires_at'] <= time.time():
            print('Access token expired, requesting new one.')
            token = get_token(client_id, client_secret, auth_url)
            save_token(token_file_name, token)
        else:
            print('Using existing access token.')

    return token


def get_request(token_object, url):
    """use token to send GET request to the API url and return the response in JSON format"""
    token = token_object['access_token']
    token_type = token_object['token_type']
    authentication = {'Authorization': '{} {}'.format(token_type, token), 'Accept': 'application/json',
                      'Host': 'api.cisco.com'}

    try:
        response = requests.get(url, headers=authentication)
    except requests.ConnectionError:
        print('Error connecting to {}, continuing to next product.'.format(url))
        return False

    # return False if the GET request is unsuccessful
    try:
        response.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print(str(e))
        return False
    except requests.RequestException:
        print(response.status_code)
        return False
    else:
        return response.json()


def eox_value(entry, list):
    """extract the required dictionary value from the EOX list"""
    for entries in list:
        try:
            item = entries[entry]
        except KeyError:
            return False
        else:
            return item['value']  # the resulting item is itself a dictionary, return the value in the 'value' key


def format_date(date_str):
    """takes a string, converts to a date object, reformats the date, then returns the new format date as a string"""
    try:
        date_object = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return date_str
    else:
        formatted_str = date_object.strftime('%d/%m/%Y')
        return formatted_str


def open_workbook(name):
    """open excel workbook and handle exceptions"""
    print('Opening file: {}'.format(name))
    try:
        wb = openpyxl.load_workbook(name)
    except IOError:
        print('Could not open file: {}'.format(name))
        input('Press enter to continue')
        sys.exit()
    else:
        return wb


def find_column(sheet, title):
    """search the first row of a spreadsheet for a column title, return column index of first match"""
    for col in sheet.iter_cols(max_row=1):    # return tuple of values for 1st row of each column
        for cell in col:                              # loop through entries in column tuple
            if cell.value == title:
                assert isinstance(cell.col_idx, int)  # column index must be an integer
                match = cell.col_idx
                return match

    # if a match is not found, return an error and stop the program
    print('Column \'{}\' not found in \'{}\''.format(title, sheet))
    input('Press enter to continue')
    sys.exit()


def find_rows(column, product_id):
    """search a column for matching device ID's and return list of row numbers"""
    matches = []
    for cell in column:
        if str(cell.value).casefold() == product_id.casefold():  # .casefold() ignores upper and lowercase differences
            matches.append(cell.row)
    return matches


def hw_process(current_sheet):
    """This function executes the API call and storage of the retrieved data for the Hardware.

     The API call is made and the end-of-life data extracted from the response. The relevant cells in the workbook are
     then updated and timestamped, and a report on the changes made is saved to a text file.
     """

    print('Updating {}'.format(current_sheet))
    updated_products = []  # list to record the updated product ID's
    part_column_number = find_column(current_sheet, 'PartNumber')  # get number of column titled 'PartNumber'
    column_values = list(current_sheet.columns)[part_column_number - 1]  # get list of cells in the
    # 'PartNumber' column
    ldos_column_number = find_column(current_sheet, 'HWEndofSupportDate')
    eol_column_number = find_column(current_sheet, 'HWEndofLifeDate')

    # iterate through the list of cells in column 'PartNumber'
    count = 0
    for part_number in column_values:
        if part_number.row is not 1:  # do not change 1st (title) row
            if count >= scan_limit:
                break

            product_id = part_number.value
            # skip if product_id cell is blank
            if product_id:
                print(product_id)
            else:
                continue

            count += 1

            # create API url by appending 'part_number' to the base url
            url = '{}/{}'.format(hw_base_url, product_id)

            # send GET request to API, get_request function will return False if unsuccessful
            product_data = get_request(token, url)
            if product_data is False:
                continue

            # extract EOX section from response, skip if not present
            try:
                eox_data = product_data['EOXRecord']
            except KeyError:
                continue

            # check for error and skip product if error is present
            error_present = False
            for response_dict in eox_data:
                for entry in response_dict.keys():
                    if entry == 'EOXError':
                        # error_message = response_dict['EOXError']['ErrorDescription']
                        error_present = True
                        # print(error_message)
                        break
            if error_present is True:
                continue

            # extract required information from the EOX section
            ldos_date = format_date(eox_value('LastDateOfSupport', eox_data))
            eol_date = format_date(eox_value('EndOfSWMaintenanceReleases', eox_data))

            # update HWEndofSupportDate cell if API value is not blank
            if ldos_date:
                current_sheet.cell(row=part_number.row, column=ldos_column_number).value = ldos_date
            else:
                if not ldos_date:
                    ldos_date = 'Not available'

            # update HWEndofLifeDate cell if API value is not blank
            if eol_date:
                current_sheet.cell(row=part_number.row, column=eol_column_number).value = eol_date
            else:
                if not eol_date:
                    eol_date = False

            # timestamp the 'Updated' column
            if not (ldos_date == False and eol_date == False):
                date_column_number = find_column(current_sheet, 'Updated')
                timestamp = datetime.now()
                current_sheet.cell(row=part_number.row, column=date_column_number).value = timestamp

            # create record of product and updated info
            update = (product_id, ldos_date, eol_date)

            # add this product to list of updated products
            updated_products.append(update)

    print('HW Products updated: {}'.format(len(updated_products)))
    # create the report of results
    create_report(current_sheet, count, len(updated_products),
                  updated_products, 'HW Report {}'.format(date.today()))

    return


def sw_process(current_sheet):
    """This function executes the API call and storage of the retrieved data for the Software.

    The API call is made and the end-of-life data extracted from the response. The relevant cells in the workbook are
    then updated and timestamped, and a report on the changes made is saved to a text file.
    """

    print('Updating {}'.format(current_sheet))
    updated_products = []  # list to record the updated product ID's
    software_column_number = find_column(current_sheet, 'oslevel')  # get number of column titled 'oslevel'
    column_values = list(current_sheet.columns)[
        software_column_number - 1]  # get list of cells in the 'oslevel' column
    ldos_column_number = find_column(current_sheet,
                                     'SWEndofSupportDate')  # get number of column titled 'SWEndOfSupportDate'

    # iterate through the list of cells in column 'oslevel'
    count = 0
    for software in column_values:
        if software.row is not 1:  # do not change 1st (title) row
            if count >= scan_limit:
                break

            software_id = software.value
            # skip if software_id cell is blank
            if software_id:
                print(software_id)
            else:
                continue

            count += 1

            # create API url by appending 'software_id' to the base url
            url = '{}/?input={}'.format(sw_base_url, str(software_id))

            # send GET request to API, get_request function will return False if unsuccessful
            product_data = get_request(token, url)
            if product_data is False:
                print('Error retrieving data for {}'.format(software_id))
                continue

            # extract EOX section from response, skip if not present
            try:
                eox = product_data['EOXRecord']
            except KeyError:
                continue

            # check for error and skip product if error is present
            error_present = False
            for response_dict in eox:
                for entry in response_dict.keys():
                    if entry == 'EOXError':
                        # error_message = response_dict['EOXError']['ErrorDescription']
                        error_present = True
                        # print(error_message)
                        break
            if error_present is True:
                continue

            # extract required information from the EOX section
            ldos_date = format_date(eox_value('LastDateOfSupport', eox))

            # update SWEndofSupportDate cell if API value is not blank
            if ldos_date:
                current_sheet.cell(row=software.row, column=ldos_column_number).value = ldos_date
            else:
                if not ldos_date:
                    ldos_date = False

            # timestamp the 'Updated' column
            if not ldos_date == False:
                date_column_number = find_column(current_sheet, 'Updated')
                date_now = datetime.now()
                current_sheet.cell(row=software.row, column=date_column_number).value = date_now

            # create record of product and updated info
            update = (software_id, ldos_date)

            # add this product to list of updated products
            updated_products.append(update)

    print('SW Products updated: {}'.format(len(updated_products)))
    # create the report of results
    create_report(current_sheet, count, len(updated_products),
                  updated_products, 'SW Report {}'.format(date.today()))

    return


begin = time.time()

token = check_token(token_file_name)  # create the access token
wb = open_workbook(excel_name)

for sheet in wb.sheetnames:
    current_sheet = wb[sheet]

    # find the sheet named 'HW Product Catalogue'
    if sheet == hw_sheet_name:
        hw_process(current_sheet)

    # find the sheet named 'SW Product Catalogue'
    if sheet == sw_sheet_name:
        sw_process(current_sheet)


save_as = 'Product Catalogue Updated {}.xlsx'.format(date.today())  # put current date in the updated file name
try:
    wb.save(save_as)
    print('Update complete, saved file \'{}\''.format(save_as))
except PermissionError as e:
    print('Saving file {} : {}\n'
          'File of same name may already be open.'.format(save_as, e.strerror))
    input('Press enter to continue')

end = time.time()
runtime = end - begin
runtime_mins = runtime/60

print('Runtime = {} minute(s)'.format(round(runtime_mins)))
input('Press enter to continue')



